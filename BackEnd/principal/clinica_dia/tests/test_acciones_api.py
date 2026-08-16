from datetime import timedelta
from datetime import datetime, time
from io import BytesIO
from unittest.mock import patch

from django.urls import reverse
from django.utils import timezone
from openpyxl import load_workbook
from rest_framework import status
from rest_framework.authtoken.models import Token
from rest_framework.test import APITestCase

from clinica_dia.models import ProgramacionQuimioterapia, SolicitudQuimioterapia
from clinica_dia.services import siguiente_dia_habil
from usuarios.models import Usuario


class AccionesClinicaDiaAPITests(APITestCase):
    def setUp(self):
        self.admin = Usuario.objects.create_user(
            username="admin-acciones-clinica",
            email="admin.acciones.clinica@example.com",
            password="Clave-segura-2026",
            rol=Usuario.Rol.ADMINISTRADOR,
            estado=Usuario.Estado.ACTIVO,
            is_active=True,
            is_staff=True,
        )
        self.medico = Usuario.objects.create_user(
            username="medico-sin-acceso-clinica",
            email="medico.sin.acceso.clinica@example.com",
            password="Clave-segura-2026",
            rol=Usuario.Rol.MEDICO,
            estado=Usuario.Estado.ACTIVO,
            is_active=True,
        )
        self.fecha = siguiente_dia_habil(timezone.localdate() + timedelta(days=1))
        self._autenticar(self.admin)

    def _autenticar(self, usuario):
        token, _ = Token.objects.get_or_create(user=usuario)
        self.client.credentials(HTTP_AUTHORIZATION=f"Bearer {token.key}")

    def _solicitud(self, codigo, *, procedencia="Hematología"):
        return SolicitudQuimioterapia.objects.create(
            codigo_externo=codigo,
            dni="71234567",
            nombre_completo_importado=f"Paciente {codigo}",
            telefono="987654321",
            procedencia=procedencia,
            prioridad=SolicitudQuimioterapia.Prioridad.ALTA,
            fecha_preferida=self.fecha,
            duracion_minutos=60,
            origen=SolicitudQuimioterapia.Origen.MANUAL,
            creada_por=self.admin,
        )

    def _programar(self, solicitud, cama=1, recordatorio=True):
        return self.client.post(
            reverse("clinica_dia:programar"),
            {
                "solicitud_id": str(solicitud.id),
                "fecha": self.fecha.isoformat(),
                "turno": "T1",
                "cama": cama,
                "crear_recordatorio": recordatorio,
            },
            format="json",
        )

    def test_solo_administrador_accede_al_panel(self):
        self._autenticar(self.medico)
        respuesta = self.client.get(reverse("clinica_dia:panel"), {"fecha": self.fecha})
        self.assertEqual(respuesta.status_code, status.HTTP_403_FORBIDDEN)

    def test_panel_programar_colision_confirmar_agenda_recordatorio_y_completar(self):
        solicitud = self._solicitud("ACC-001", procedencia="Hospitalización")
        programada = self._programar(solicitud, recordatorio=False)
        self.assertEqual(programada.status_code, status.HTTP_201_CREATED, programada.data)
        programacion_id = programada.data["programacion"]["id"]
        self.assertEqual(
            programada.data["programacion"]["recordatorio_estado"], "NO_REQUERIDO"
        )
        segunda = self._solicitud("ACC-002")
        colision = self._programar(segunda)
        self.assertEqual(colision.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertEqual(self._programar(segunda, cama=2).status_code, status.HTTP_201_CREATED)

        recordatorio = self.client.patch(
            reverse("clinica_dia:recordatorio-programacion", args=[programacion_id]),
            {"estado": "ENVIADO", "observacion": "Tutor informado."},
            format="json",
        )
        self.assertEqual(recordatorio.status_code, status.HTTP_200_OK)
        self.assertEqual(recordatorio.data["programacion"]["recordatorio_estado"], "ENVIADO")

        panel = self.client.get(
            reverse("clinica_dia:panel"),
            {"fecha": self.fecha.isoformat(), "procedencia": "Hospitalización"},
        )
        self.assertEqual(panel.status_code, status.HTTP_200_OK)
        self.assertEqual(panel.data["resumen"]["camas_ocupadas"], 2)
        self.assertIn("Hospitalización", panel.data["procedencias"])

        confirmacion = self.client.post(
            reverse("clinica_dia:confirmar-agenda"),
            {"fecha": self.fecha.isoformat()},
            format="json",
        )
        self.assertEqual(confirmacion.data["confirmadas"], 2)
        self.assertEqual(confirmacion.data["total"], 2)
        completar = self.client.post(
            reverse("clinica_dia:completar-programacion", args=[programacion_id]),
            {},
            format="json",
        )
        self.assertEqual(completar.status_code, status.HTTP_400_BAD_REQUEST)
        instante_fin = timezone.make_aware(datetime.combine(self.fecha, time(18, 0)))
        with patch("clinica_dia.services.timezone.localtime", return_value=instante_fin):
            completar = self.client.post(
                reverse("clinica_dia:completar-programacion", args=[programacion_id]),
                {},
                format="json",
            )
        self.assertEqual(completar.status_code, status.HTTP_200_OK)
        self.assertEqual(completar.data["programacion"]["estado"], "COMPLETADA")
        solicitud.refresh_from_db()
        self.assertEqual(solicitud.estado, SolicitudQuimioterapia.Estado.COMPLETADA)

    def test_cancelar_reencola_y_exportar_excluye_cancelada_y_neutraliza_formula(self):
        solicitud = self._solicitud("ACC-CSV", procedencia="=FORMULA")
        programada = self._programar(solicitud)
        programacion_id = programada.data["programacion"]["id"]
        cancelada = self.client.post(
            reverse("clinica_dia:cancelar-programacion", args=[programacion_id]),
            {"motivo": "Paciente solicitó nueva fecha.", "reprogramar": True},
            format="json",
        )
        self.assertEqual(cancelada.status_code, status.HTTP_200_OK)
        solicitud.refresh_from_db()
        self.assertEqual(solicitud.estado, SolicitudQuimioterapia.Estado.PENDIENTE)
        exportacion_vacia = self.client.get(
            reverse("clinica_dia:exportar"),
            {"fecha_desde": self.fecha.isoformat(), "formato": "xlsx"},
        )
        libro = load_workbook(BytesIO(exportacion_vacia.content), data_only=False)
        self.assertEqual(libro["Agenda"].max_row, 1)
        libro.close()

        activa = self._solicitud("ACC-CSV-2", procedencia="=FORMULA")
        self._programar(activa, cama=2)
        exportacion = self.client.get(
            reverse("clinica_dia:exportar"),
            {"fecha_desde": self.fecha.isoformat(), "formato": "xlsx"},
        )
        libro = load_workbook(BytesIO(exportacion.content), data_only=False)
        self.assertEqual(libro["Agenda"].cell(row=2, column=6).value, "'=FORMULA")
        libro.close()
