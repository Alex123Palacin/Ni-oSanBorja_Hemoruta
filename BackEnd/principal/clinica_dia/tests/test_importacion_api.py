from io import BytesIO
from datetime import timedelta

from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import override_settings
from django.urls import reverse
from django.utils import timezone
from openpyxl import Workbook, load_workbook
from rest_framework import status
from rest_framework.authtoken.models import Token
from rest_framework.test import APITestCase

from clinica_dia.importacion import COLUMNAS_PERMITIDAS
from clinica_dia.models import SolicitudQuimioterapia
from clinica_dia.services import siguiente_dia_habil
from pacientes.models import Paciente
from usuarios.models import Usuario


@override_settings(MEDIA_ROOT=None)
class ImportacionClinicaDiaAPITests(APITestCase):
    def setUp(self):
        self.admin = Usuario.objects.create_user(
            username="admin-clinica-dia",
            password="Clave-segura-2026",
            rol=Usuario.Rol.ADMINISTRADOR,
            estado=Usuario.Estado.ACTIVO,
            is_active=True,
            is_staff=True,
        )
        token = Token.objects.create(user=self.admin)
        self.client.credentials(HTTP_AUTHORIZATION=f"Bearer {token.key}")
        self.fecha = siguiente_dia_habil(timezone.localdate() + timedelta(days=1))

    def _xlsx(self, filas, *, hoja="Pacientes"):
        libro = Workbook()
        pagina = libro.active
        pagina.title = hoja
        pagina.append(COLUMNAS_PERMITIDAS)
        for fila in filas:
            pagina.append(fila)
        contenido = BytesIO()
        libro.save(contenido)
        libro.close()
        return SimpleUploadedFile(
            "solicitudes.xlsx",
            contenido.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    def _fila(self, codigo="SOL-001", dni="12345678"):
        return [
            codigo,
            dni,
            "Mateo Gabriel Flores",
            "987654321",
            self.fecha,
            "11:30",
            2,
            "ALTA",
            "Hematología",
            "HC-IMPORT-001",
            "Leucemia",
            "",
            "Programar con acompañante.",
        ]

    def test_importa_en_memoria_vincula_por_dni_y_repeticion_es_duplicado(self):
        paciente = Paciente.objects.create(
            historia_clinica="HC-IMPORT-001",
            dni="12345678",
            nombres="Mateo Gabriel",
            apellidos="Flores",
        )
        respuesta = self.client.post(
            reverse("clinica_dia:importar"),
            {"archivo": self._xlsx([self._fila()])},
            format="multipart",
        )
        self.assertEqual(respuesta.status_code, status.HTTP_201_CREATED, respuesta.data)
        self.assertEqual(respuesta.data["total"], 1)
        self.assertEqual(respuesta.data["importadas"], 1)
        self.assertEqual(respuesta.data["vinculadas"], 1)
        self.assertEqual(respuesta.data["rechazadas"], 0)
        solicitud = SolicitudQuimioterapia.objects.get(codigo_externo="SOL-001")
        self.assertEqual(solicitud.paciente, paciente)
        self.assertEqual(solicitud.procedencia, "Hematología")
        self.assertEqual(solicitud.historia_clinica_importada, "HC-IMPORT-001")
        self.assertEqual(solicitud.protocolo, "")

        repetida = self.client.post(
            reverse("clinica_dia:importar"),
            {"archivo": self._xlsx([self._fila()])},
            format="multipart",
        )
        self.assertEqual(repetida.status_code, status.HTTP_200_OK)
        self.assertEqual(repetida.data["duplicadas"], 1)
        self.assertEqual(SolicitudQuimioterapia.objects.count(), 1)

    def test_reporta_errores_por_fila_y_no_guarda_filas_invalidas(self):
        fila = self._fila(codigo="   ", dni="ABC")
        fila[3] = "123"
        fila[4] = self.fecha + timedelta(days=(5 - self.fecha.weekday()) % 7)
        fila[5] = "20:00"
        fila[6] = 5
        fila[7] = "URGENTE"
        respuesta = self.client.post(
            reverse("clinica_dia:importar"),
            {"archivo": self._xlsx([fila])},
            format="multipart",
        )
        self.assertEqual(respuesta.status_code, status.HTTP_200_OK)
        self.assertEqual(respuesta.data["rechazadas"], 1)
        campos = {error["campo"] for error in respuesta.data["errores"]}
        self.assertTrue(
            {"codigo_solicitud", "dni", "telefono", "hora_preferida", "duracion_horas", "prioridad"}
            <= campos
        )
        self.assertFalse(SolicitudQuimioterapia.objects.exists())

    def test_rechaza_sin_archivo_y_mas_de_cinco_mil_filas_no_vacias(self):
        sin_archivo = self.client.post(reverse("clinica_dia:importar"), {}, format="multipart")
        self.assertEqual(sin_archivo.status_code, status.HTTP_400_BAD_REQUEST)

        filas = [self._fila(codigo=f"SOL-{indice:05d}", dni=f"{indice:08d}") for indice in range(1, 5002)]
        demasiadas = self.client.post(
            reverse("clinica_dia:importar"),
            {"archivo": self._xlsx(filas)},
            format="multipart",
        )
        self.assertEqual(demasiadas.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn("5000", str(demasiadas.data))
        self.assertFalse(SolicitudQuimioterapia.objects.exists())

    def test_plantilla_tiene_esquema_exacto_sin_fila_ficticia(self):
        respuesta = self.client.get(reverse("clinica_dia:plantilla"))
        self.assertEqual(respuesta.status_code, status.HTTP_200_OK)
        self.assertIn("plantilla_clinica_dia.xlsx", respuesta["Content-Disposition"])
        libro = load_workbook(BytesIO(respuesta.content), read_only=True, data_only=True)
        hoja = libro["Pacientes"]
        self.assertEqual(tuple(celda.value for celda in hoja[1]), COLUMNAS_PERMITIDAS)
        self.assertFalse(any(celda.value is not None for celda in hoja[2]))
        self.assertIn("Instrucciones", libro.sheetnames)
        libro.close()
