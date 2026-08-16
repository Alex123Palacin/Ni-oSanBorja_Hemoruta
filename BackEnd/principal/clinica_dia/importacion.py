from __future__ import annotations

import hashlib
import json
import re
import uuid
from datetime import date, datetime, time, timedelta
from decimal import Decimal, InvalidOperation
from io import BytesIO
from zipfile import BadZipFile, ZipFile

from django.core.exceptions import ValidationError
from django.db import IntegrityError, transaction
from django.utils import timezone
from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel

from pacientes.models import Paciente

from .models import SolicitudQuimioterapia
from .services import es_dia_habil


HOJA_SOLICITUDES = "Pacientes"
COLUMNAS_OBLIGATORIAS = (
    "codigo_solicitud",
    "dni",
    "nombres_completos",
    "telefono",
    "fecha_solicitada",
    "hora_preferida",
    "duracion_horas",
    "prioridad",
    "procedencia",
)
COLUMNAS_OPCIONALES = (
    "historia_clinica",
    "diagnostico",
    "protocolo_quimioterapia",
    "observaciones",
)
COLUMNAS_PERMITIDAS = COLUMNAS_OBLIGATORIAS + COLUMNAS_OPCIONALES
MAXIMO_FILAS = 5000


def _texto(valor) -> str:
    if valor is None:
        return ""
    if isinstance(valor, float) and valor.is_integer():
        return str(int(valor))
    return str(valor).strip()


def _dni(valor) -> str:
    texto = _texto(valor)
    if re.fullmatch(r"\d+(?:\.0+)?", texto):
        texto = str(int(float(texto))).zfill(8)
    if not re.fullmatch(r"\d{8}", texto):
        raise ValueError("Debe contener exactamente 8 dígitos.")
    return texto


def _telefono(valor) -> str:
    texto = re.sub(r"\D", "", _texto(valor))
    if not 9 <= len(texto) <= 12:
        raise ValueError("Debe contener entre 9 y 12 dígitos.")
    return texto


def _fecha(valor, epoch) -> date:
    if isinstance(valor, datetime):
        resultado = valor.date()
    elif isinstance(valor, date):
        resultado = valor
    elif isinstance(valor, (int, float)):
        convertido = from_excel(valor, epoch)
        resultado = convertido.date() if isinstance(convertido, datetime) else convertido
    else:
        texto = _texto(valor)
        resultado = None
        for formato in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
            try:
                resultado = datetime.strptime(texto, formato).date()
                break
            except ValueError:
                continue
        if resultado is None:
            raise ValueError("Use una fecha válida (AAAA-MM-DD o DD/MM/AAAA).")
    if not es_dia_habil(resultado):
        raise ValueError("La fecha debe ser un día hábil, de lunes a viernes.")
    if resultado < timezone.localdate():
        raise ValueError("La fecha solicitada no puede estar en el pasado.")
    return resultado


def _hora(valor) -> time:
    if isinstance(valor, datetime):
        resultado = valor.time().replace(second=0, microsecond=0)
    elif isinstance(valor, time):
        resultado = valor.replace(second=0, microsecond=0)
    elif isinstance(valor, (int, float)):
        segundos = round((float(valor) % 1) * 24 * 60 * 60)
        resultado = (datetime.min + timedelta(seconds=segundos)).time().replace(second=0)
    else:
        texto = _texto(valor)
        resultado = None
        for formato in ("%H:%M", "%H:%M:%S", "%I:%M %p"):
            try:
                resultado = datetime.strptime(texto.upper(), formato).time()
                break
            except ValueError:
                continue
        if resultado is None:
            raise ValueError("Use una hora válida en formato HH:MM.")
    if not (time(8, 0) <= resultado < time(17, 30)):
        raise ValueError("Debe estar entre las 08:00 y las 17:29.")
    return resultado


def _duracion_minutos(valor) -> int:
    try:
        horas = Decimal(_texto(valor).replace(",", "."))
    except (InvalidOperation, ValueError):
        raise ValueError("Debe ser un número de horas válido.") from None
    minutos_decimal = horas * 60
    minutos = int(minutos_decimal)
    if minutos_decimal != minutos or not 15 <= minutos <= 210:
        raise ValueError("Debe equivaler a una cantidad exacta entre 15 y 210 minutos.")
    return minutos


def _prioridad(valor) -> str:
    resultado = _texto(valor).upper()
    if resultado not in SolicitudQuimioterapia.Prioridad.values:
        raise ValueError("Use ALTA, MEDIA o BAJA.")
    return resultado


def _error(fila: int, campo: str, mensaje: str) -> dict:
    return {"fila": fila, "campo": campo, "mensaje": mensaje}


def _validar_fila(datos: dict, fila: int, epoch) -> tuple[dict | None, list[dict]]:
    errores: list[dict] = []
    resultado: dict = {}
    for campo in COLUMNAS_OBLIGATORIAS:
        if not _texto(datos.get(campo)):
            errores.append(_error(fila, campo, "Este valor es obligatorio."))
    validadores = {
        "dni": _dni,
        "telefono": _telefono,
        "hora_preferida": _hora,
        "duracion_horas": _duracion_minutos,
        "prioridad": _prioridad,
    }
    for campo, validador in validadores.items():
        if datos.get(campo) in (None, ""):
            continue
        try:
            resultado[campo] = validador(datos[campo])
        except ValueError as error:
            errores.append(_error(fila, campo, str(error)))
    if datos.get("fecha_solicitada") not in (None, ""):
        try:
            resultado["fecha_solicitada"] = _fecha(datos["fecha_solicitada"], epoch)
        except ValueError as error:
            errores.append(_error(fila, "fecha_solicitada", str(error)))
    for campo in (
        "codigo_solicitud",
        "nombres_completos",
        "procedencia",
        "historia_clinica",
        "diagnostico",
        "protocolo_quimioterapia",
        "observaciones",
    ):
        resultado[campo] = _texto(datos.get(campo))
    if len(resultado["codigo_solicitud"]) > 80:
        errores.append(_error(fila, "codigo_solicitud", "No debe superar 80 caracteres."))
    if len(resultado["nombres_completos"]) > 241:
        errores.append(_error(fila, "nombres_completos", "No debe superar 241 caracteres."))
    if len(resultado["procedencia"]) > 160:
        errores.append(_error(fila, "procedencia", "No debe superar 160 caracteres."))
    if len(resultado["historia_clinica"]) > 32:
        errores.append(_error(fila, "historia_clinica", "No debe superar 32 caracteres."))
    if len(resultado["diagnostico"]) > 240:
        errores.append(_error(fila, "diagnostico", "No debe superar 240 caracteres."))
    if len(resultado["protocolo_quimioterapia"]) > 240:
        errores.append(
            _error(fila, "protocolo_quimioterapia", "No debe superar 240 caracteres.")
        )
    return (None if errores else resultado), errores


def _huella(datos: dict) -> str:
    serializable = {
        clave: valor.isoformat() if isinstance(valor, (date, time)) else valor
        for clave, valor in datos.items()
    }
    contenido = json.dumps(serializable, sort_keys=True, ensure_ascii=False).encode("utf-8")
    return hashlib.sha256(contenido).hexdigest()


def _resolver_paciente(datos: dict, fila: int) -> tuple[Paciente | None, list[dict]]:
    paciente_dni = Paciente.objects.filter(dni=datos["dni"]).first()
    historia = datos["historia_clinica"]
    paciente_historia = (
        Paciente.objects.filter(historia_clinica=historia).first() if historia else None
    )
    if paciente_dni and historia and paciente_dni.historia_clinica != historia:
        return None, [_error(fila, "historia_clinica", "No coincide con el DNI registrado.")]
    if paciente_historia and paciente_historia.dni and paciente_historia.dni != datos["dni"]:
        return None, [_error(fila, "dni", "No coincide con la historia clínica registrada.")]
    return paciente_dni or paciente_historia, []


def importar_solicitudes_xlsx(archivo, *, usuario) -> dict:
    try:
        contenido = archivo.read()
        with ZipFile(BytesIO(contenido)) as comprimido:
            if len(comprimido.infolist()) > 1000 or sum(
                entrada.file_size for entrada in comprimido.infolist()
            ) > 50 * 1024 * 1024:
                raise ValidationError({"archivo": "El contenido descomprimido es demasiado grande."})
        libro = load_workbook(
            BytesIO(contenido), read_only=True, data_only=True, keep_links=False
        )
    except ValidationError:
        raise
    except BadZipFile as error:
        raise ValidationError({"archivo": "El XLSX está dañado o no es válido."}) from error
    except Exception as error:
        raise ValidationError({"archivo": "El XLSX está dañado o no es válido."}) from error
    try:
        if HOJA_SOLICITUDES not in libro.sheetnames:
            raise ValidationError(
                {"archivo": f'El archivo debe contener la hoja "{HOJA_SOLICITUDES}".'}
            )
        hoja = libro[HOJA_SOLICITUDES]
        if hoja.max_row > 100000:
            raise ValidationError({"archivo": "La hoja declara demasiadas filas."})
        encabezados = [_texto(celda.value).lower() for celda in hoja[1]]
        faltantes = [campo for campo in COLUMNAS_OBLIGATORIAS if campo not in encabezados]
        desconocidas = [campo for campo in encabezados if campo and campo not in COLUMNAS_PERMITIDAS]
        repetidas = sorted({campo for campo in encabezados if campo and encabezados.count(campo) > 1})
        if faltantes or desconocidas or repetidas:
            mensajes = []
            if faltantes:
                mensajes.append(f"Faltan columnas: {', '.join(faltantes)}.")
            if desconocidas:
                mensajes.append(f"Columnas no reconocidas: {', '.join(desconocidas)}.")
            if repetidas:
                mensajes.append(f"Columnas repetidas: {', '.join(repetidas)}.")
            raise ValidationError({"archivo": " ".join(mensajes)})

        posiciones = {campo: encabezados.index(campo) for campo in COLUMNAS_PERMITIDAS if campo in encabezados}
        filas_no_vacias = []
        for numero_fila, valores in enumerate(
            hoja.iter_rows(min_row=2, values_only=True), start=2
        ):
            if not any(valor not in (None, "") for valor in valores):
                continue
            filas_no_vacias.append((numero_fila, valores))
            if len(filas_no_vacias) > MAXIMO_FILAS:
                raise ValidationError(
                    {"archivo": f"El archivo admite como máximo {MAXIMO_FILAS} filas no vacías."}
                )
        lote_id = uuid.uuid4()
        resultado = {
            "detalle": "Importación finalizada.",
            "lote_id": str(lote_id),
            "total": 0,
            "importadas": 0,
            "vinculadas": 0,
            "externas": 0,
            "duplicadas": 0,
            "rechazadas": 0,
            "errores": [],
        }
        for numero_fila, valores in filas_no_vacias:
            resultado["total"] += 1
            fila = {
                campo: valores[indice] if indice < len(valores) else None
                for campo, indice in posiciones.items()
            }
            datos, errores = _validar_fila(fila, numero_fila, libro.epoch)
            if datos:
                paciente, errores_paciente = _resolver_paciente(datos, numero_fila)
                errores.extend(errores_paciente)
            else:
                paciente = None
            if errores:
                resultado["rechazadas"] += 1
                resultado["errores"].extend(errores)
                continue

            huella = _huella(datos)
            existente_codigo = SolicitudQuimioterapia.objects.filter(
                codigo_externo=datos["codigo_solicitud"]
            ).first()
            existente_huella = SolicitudQuimioterapia.objects.filter(
                huella_importacion=huella
            ).first()
            if existente_codigo or existente_huella:
                if existente_codigo and existente_codigo.huella_importacion != huella:
                    resultado["rechazadas"] += 1
                    resultado["errores"].append(
                        _error(
                            numero_fila,
                            "codigo_solicitud",
                            "Ya existe con información diferente.",
                        )
                    )
                else:
                    resultado["duplicadas"] += 1
                continue
            try:
                with transaction.atomic():
                    solicitud = SolicitudQuimioterapia(
                        paciente=paciente,
                        dni=datos["dni"],
                        nombre_completo_importado=datos["nombres_completos"],
                        historia_clinica_importada=datos["historia_clinica"],
                        telefono=datos["telefono"],
                        procedencia=datos["procedencia"],
                        diagnostico=datos["diagnostico"],
                        protocolo=datos["protocolo_quimioterapia"],
                        prioridad=datos["prioridad"],
                        fecha_preferida=datos["fecha_solicitada"],
                        hora_preferida=datos["hora_preferida"],
                        duracion_minutos=datos["duracion_horas"],
                        origen=SolicitudQuimioterapia.Origen.IMPORTACION,
                        codigo_externo=datos["codigo_solicitud"],
                        lote_importacion=lote_id,
                        fila_importacion=numero_fila,
                        huella_importacion=huella,
                        observaciones=datos["observaciones"],
                        creada_por=usuario,
                    )
                    solicitud.full_clean()
                    solicitud.save()
            except (IntegrityError, ValidationError) as error:
                resultado["rechazadas"] += 1
                mensaje = getattr(error, "message", None) or "No se pudo registrar la fila."
                resultado["errores"].append(_error(numero_fila, "fila", str(mensaje)))
                continue
            resultado["importadas"] += 1
            resultado["vinculadas" if paciente else "externas"] += 1
        resultado["detalle"] = (
            f"Importación finalizada: {resultado['importadas']} registradas, "
            f"{resultado['duplicadas']} duplicadas y {resultado['rechazadas']} rechazadas."
        )
        return resultado
    finally:
        libro.close()
