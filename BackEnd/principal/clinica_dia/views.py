from __future__ import annotations

import csv
from datetime import datetime, timedelta
from io import BytesIO, StringIO

from django.core.exceptions import ValidationError as DjangoValidationError
from django.db.models import Case, IntegerField, Q, Value, When
from django.http import HttpResponse
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from rest_framework import status
from rest_framework.exceptions import ValidationError
from rest_framework.parsers import FormParser, MultiPartParser
from rest_framework.response import Response
from rest_framework.views import APIView

from usuarios.permissions import EsAdministrador

from .importacion import COLUMNAS_PERMITIDAS, HOJA_SOLICITUDES, importar_solicitudes_xlsx
from .models import (
    HORARIOS_TURNO,
    CambioProgramacionQuimioterapia,
    ProgramacionQuimioterapia,
    SolicitudQuimioterapia,
)
from .serializers import (
    AjustarProgramacionSerializer,
    CancelarProgramacionSerializer,
    ConfirmarAgendaSerializer,
    GenerarAgendaSerializer,
    ImportarSolicitudesSerializer,
    ProgramacionQuimioterapiaSerializer,
    ProgramarSolicitudSerializer,
    RecordatorioSerializer,
    SolicitudQuimioterapiaCrearSerializer,
    SolicitudQuimioterapiaSerializer,
)
from .services import (
    actualizar_recordatorio,
    ajustar_programacion,
    cancelar_programacion,
    completar_programacion,
    confirmar_agenda,
    confirmar_programacion,
    crear_programacion,
    es_dia_habil,
    generar_agenda_automatica,
    siguiente_dia_habil,
)


def _error_drf(error: DjangoValidationError) -> ValidationError:
    if hasattr(error, "message_dict"):
        return ValidationError(error.message_dict)
    return ValidationError(error.messages)


def _programaciones_base():
    return ProgramacionQuimioterapia.objects.select_related(
        "solicitud", "solicitud__paciente", "cita"
    )


class ClinicaDiaPanelAPIView(APIView):
    permission_classes = (EsAdministrador,)

    def get(self, request):
        valor_fecha = request.query_params.get("fecha")
        try:
            fecha = (
                datetime.strptime(valor_fecha, "%Y-%m-%d").date()
                if valor_fecha
                else siguiente_dia_habil(timezone.localdate())
            )
        except ValueError as error:
            raise ValidationError({"fecha": "Use el formato AAAA-MM-DD."}) from error

        pendientes_global = SolicitudQuimioterapia.objects.filter(
            estado=SolicitudQuimioterapia.Estado.PENDIENTE
        )
        solicitudes = SolicitudQuimioterapia.objects.select_related("paciente")
        estado = request.query_params.get("estado")
        if estado and estado not in SolicitudQuimioterapia.Estado.values:
            raise ValidationError(
                {"estado": f"Use uno de: {', '.join(SolicitudQuimioterapia.Estado.values)}."}
            )
        solicitudes = solicitudes.filter(
            estado=estado or SolicitudQuimioterapia.Estado.PENDIENTE
        )
        prioridad = request.query_params.get("prioridad")
        if prioridad:
            if prioridad not in SolicitudQuimioterapia.Prioridad.values:
                raise ValidationError({"prioridad": "Use ALTA, MEDIA o BAJA."})
            solicitudes = solicitudes.filter(prioridad=prioridad)
        procedencia = request.query_params.get("procedencia", "").strip()
        if procedencia:
            solicitudes = solicitudes.filter(procedencia__iexact=procedencia)
        busqueda = request.query_params.get("q", "").strip()
        if busqueda:
            solicitudes = solicitudes.filter(
                Q(dni__icontains=busqueda)
                | Q(nombre_completo_importado__icontains=busqueda)
                | Q(paciente__nombres__icontains=busqueda)
                | Q(paciente__apellidos__icontains=busqueda)
                | Q(paciente__historia_clinica__icontains=busqueda)
                | Q(historia_clinica_importada__icontains=busqueda)
                | Q(codigo_externo__icontains=busqueda)
                | Q(diagnostico__icontains=busqueda)
                | Q(procedencia__icontains=busqueda)
            )
        solicitudes = solicitudes.annotate(
            orden_prioridad=Case(
                When(prioridad=SolicitudQuimioterapia.Prioridad.ALTA, then=Value(0)),
                When(prioridad=SolicitudQuimioterapia.Prioridad.MEDIA, then=Value(1)),
                default=Value(2),
                output_field=IntegerField(),
            )
        )
        pendientes = list(
            solicitudes.order_by("orden_prioridad", "fecha_preferida", "creado_en")[:200]
        )

        agenda = list(
            _programaciones_base()
            .filter(fecha=fecha)
            .exclude(estado=ProgramacionQuimioterapia.Estado.CANCELADA)
            .order_by("turno", "cama")
        )
        por_slot = {(item.turno, item.cama): item for item in agenda}
        serializador_programacion = ProgramacionQuimioterapiaSerializer
        turnos = []
        for codigo, (hora_inicio, hora_fin, _) in HORARIOS_TURNO.items():
            camas = [
                {
                    "numero": numero,
                    "programacion": (
                        serializador_programacion(por_slot[(codigo, numero)]).data
                        if (codigo, numero) in por_slot
                        else None
                    ),
                }
                for numero in range(1, 9)
            ]
            ocupadas = sum(cama["programacion"] is not None for cama in camas)
            turnos.append(
                {
                    "codigo": codigo,
                    "etiqueta": dict(ProgramacionQuimioterapia.Turno.choices)[codigo],
                    "hora_inicio": hora_inicio.strftime("%H:%M"),
                    "hora_fin": hora_fin.strftime("%H:%M"),
                    "capacidad": 8,
                    "ocupadas": ocupadas,
                    "disponibles": 8 - ocupadas,
                    "camas": camas,
                }
            )
        ocupadas = len(agenda)
        recordatorios = [
            item
            for item in agenda
            if item.recordatorio_estado
            == ProgramacionQuimioterapia.EstadoRecordatorio.PENDIENTE
            and item.estado
            in (
                ProgramacionQuimioterapia.Estado.PROGRAMADA,
                ProgramacionQuimioterapia.Estado.CONFIRMADA,
            )
        ]
        return Response(
            {
                "fecha": fecha.isoformat(),
                "resumen": {
                    "solicitudes_pendientes": pendientes_global.count(),
                    "programadas_fecha": ocupadas,
                    "confirmadas_fecha": sum(
                        item.estado == ProgramacionQuimioterapia.Estado.CONFIRMADA
                        for item in agenda
                    ),
                    "recordatorios_pendientes": len(recordatorios),
                    "capacidad_total": 24,
                    "camas_ocupadas": ocupadas,
                    "camas_disponibles": 24 - ocupadas,
                    "ocupacion_porcentaje": round((ocupadas / 24) * 100),
                },
                "turnos": turnos,
                "procedencias": list(
                    SolicitudQuimioterapia.objects.exclude(procedencia="")
                    .order_by("procedencia")
                    .values_list("procedencia", flat=True)
                    .distinct()
                ),
                "pendientes": SolicitudQuimioterapiaSerializer(pendientes, many=True).data,
                "recordatorios": serializador_programacion(recordatorios, many=True).data,
            }
        )


class SolicitudQuimioterapiaCrearAPIView(APIView):
    permission_classes = (EsAdministrador,)

    def post(self, request):
        serializer = SolicitudQuimioterapiaCrearSerializer(
            data=request.data, context={"request": request}
        )
        serializer.is_valid(raise_exception=True)
        solicitud = serializer.save()
        return Response(
            SolicitudQuimioterapiaSerializer(solicitud).data,
            status=status.HTTP_201_CREATED,
        )


class ImportarSolicitudesAPIView(APIView):
    permission_classes = (EsAdministrador,)
    parser_classes = (MultiPartParser, FormParser)

    def post(self, request):
        serializer = ImportarSolicitudesSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        try:
            resultado = importar_solicitudes_xlsx(
                serializer.validated_data["archivo"], usuario=request.user
            )
        except DjangoValidationError as error:
            raise _error_drf(error) from error
        codigo = status.HTTP_201_CREATED if resultado["importadas"] else status.HTTP_200_OK
        return Response(resultado, status=codigo)


class PlantillaSolicitudesAPIView(APIView):
    permission_classes = (EsAdministrador,)

    def get(self, request):
        libro = Workbook()
        hoja = libro.active
        hoja.title = HOJA_SOLICITUDES
        hoja.append(list(COLUMNAS_PERMITIDAS))
        ejemplo_fecha = siguiente_dia_habil(timezone.localdate() + timedelta(days=1))
        for fila in range(2, 5002):
            for indice in (1, 2, 4):
                hoja.cell(row=fila, column=indice).number_format = "@"
        for celda in hoja[1]:
            celda.font = Font(bold=True, color="FFFFFF")
            celda.fill = PatternFill("solid", fgColor="009DAA")
        hoja.freeze_panes = "A2"
        hoja.auto_filter.ref = hoja.dimensions
        for columna in hoja.columns:
            letra = columna[0].column_letter
            hoja.column_dimensions[letra].width = min(
                38, max(15, max(len(str(celda.value or "")) for celda in columna) + 2)
            )
        instrucciones = libro.create_sheet("Instrucciones")
        instrucciones.append(["Clínica de Día - importación de solicitudes"])
        instrucciones.append(["Complete una fila por sesión. No cambie los encabezados de Pacientes."])
        instrucciones.append(["Prioridad", "ALTA, MEDIA o BAJA"])
        instrucciones.append(["Fecha", f"Día hábil desde {ejemplo_fecha:%d/%m/%Y}"])
        instrucciones.append(["Hora", "Entre 08:00 y 17:29"])
        instrucciones.append(["Duración", "Entre 0.25 y 3.5 horas"])
        contenido = BytesIO()
        libro.save(contenido)
        libro.close()
        respuesta = HttpResponse(
            contenido.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        respuesta["Content-Disposition"] = 'attachment; filename="plantilla_clinica_dia.xlsx"'
        return respuesta


class GenerarAgendaAPIView(APIView):
    permission_classes = (EsAdministrador,)

    def post(self, request):
        serializer = GenerarAgendaSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        fecha_seleccionada = serializer.validated_data["fecha_desde"]
        if not es_dia_habil(fecha_seleccionada):
            raise ValidationError(
                {"fecha_desde": "Seleccione un día hábil para generar la agenda."}
            )
        # La acción del tablero agenda exclusivamente la fecha seleccionada.
        # Las solicitudes que no alcancen cama conservan su estado pendiente.
        datos_generacion = {
            **serializer.validated_data,
            "fecha_desde": fecha_seleccionada,
            "fecha_hasta": fecha_seleccionada,
        }
        try:
            creadas, no_programadas = generar_agenda_automatica(
                usuario=request.user, **datos_generacion
            )
        except DjangoValidationError as error:
            raise _error_drf(error) from error
        return Response(
            {
                "detalle": (
                    f"Agenda generada: {len(creadas)} programadas y "
                    f"{len(no_programadas)} pendientes sin cupo."
                ),
                "programadas": len(creadas),
                "no_programadas": len(no_programadas),
                "fecha_desde_efectiva": fecha_seleccionada.isoformat(),
                "errores": no_programadas,
                "programaciones": ProgramacionQuimioterapiaSerializer(creadas, many=True).data,
            },
            status=status.HTTP_201_CREATED if creadas else status.HTTP_200_OK,
        )


class ProgramacionesAPIView(APIView):
    permission_classes = (EsAdministrador,)

    def post(self, request):
        serializer = ProgramarSolicitudSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        try:
            programacion = crear_programacion(
                usuario=request.user,
                origen=ProgramacionQuimioterapia.Origen.MANUAL,
                **serializer.validated_data,
            )
        except DjangoValidationError as error:
            raise _error_drf(error) from error
        return Response(
            {
                "detalle": "Paciente programado correctamente.",
                "programacion": ProgramacionQuimioterapiaSerializer(programacion).data,
            },
            status=status.HTTP_201_CREATED,
        )


class ProgramacionDetalleAPIView(APIView):
    permission_classes = (EsAdministrador,)

    def patch(self, request, programacion_id):
        programacion = _programaciones_base().filter(pk=programacion_id).first()
        if not programacion:
            raise ValidationError({"programacion_id": "No existe la programación."})
        serializer = AjustarProgramacionSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        try:
            ajustada = ajustar_programacion(
                programacion, usuario=request.user, **serializer.validated_data
            )
        except DjangoValidationError as error:
            raise _error_drf(error) from error
        reemplazo = ajustada.pk != programacion.pk
        return Response(
            {
                "detalle": (
                    "El cupo fue reasignado y el paciente anterior volvió a pendientes."
                    if reemplazo
                    else "Programación actualizada correctamente."
                ),
                "programacion": ProgramacionQuimioterapiaSerializer(ajustada).data,
                "programacion_reemplazada_id": str(programacion.pk) if reemplazo else None,
            }
        )


class ConfirmarProgramacionAPIView(APIView):
    permission_classes = (EsAdministrador,)

    def post(self, request, programacion_id):
        programacion = _programaciones_base().filter(pk=programacion_id).first()
        if not programacion:
            raise ValidationError({"programacion_id": "No existe la programación."})
        try:
            confirmada, cambio = confirmar_programacion(programacion, usuario=request.user)
        except DjangoValidationError as error:
            raise _error_drf(error) from error
        return Response(
            {
                "detalle": "Programación confirmada." if cambio else "Ya estaba confirmada.",
                "programacion": ProgramacionQuimioterapiaSerializer(confirmada).data,
            }
        )


class ConfirmarAgendaAPIView(APIView):
    permission_classes = (EsAdministrador,)

    def post(self, request):
        serializer = ConfirmarAgendaSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        try:
            resultado = confirmar_agenda(
                fecha=serializer.validated_data["fecha"], usuario=request.user
            )
        except DjangoValidationError as error:
            raise _error_drf(error) from error
        resultado["detalle"] = (
            f"Agenda confirmada: {resultado['confirmadas']} nuevas confirmaciones."
        )
        return Response(resultado)


class CompletarProgramacionAPIView(APIView):
    permission_classes = (EsAdministrador,)

    def post(self, request, programacion_id):
        programacion = _programaciones_base().filter(pk=programacion_id).first()
        if not programacion:
            raise ValidationError({"programacion_id": "No existe la programación."})
        try:
            completada, cambio = completar_programacion(programacion, usuario=request.user)
        except DjangoValidationError as error:
            raise _error_drf(error) from error
        return Response(
            {
                "detalle": "Atención completada." if cambio else "Ya estaba completada.",
                "programacion": ProgramacionQuimioterapiaSerializer(completada).data,
            }
        )


class CancelarProgramacionAPIView(APIView):
    permission_classes = (EsAdministrador,)

    def post(self, request, programacion_id):
        programacion = _programaciones_base().filter(pk=programacion_id).first()
        if not programacion:
            raise ValidationError({"programacion_id": "No existe la programación."})
        serializer = CancelarProgramacionSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        try:
            cancelada, cambio = cancelar_programacion(
                programacion, usuario=request.user, **serializer.validated_data
            )
        except DjangoValidationError as error:
            raise _error_drf(error) from error
        return Response(
            {
                "detalle": "Programación cancelada." if cambio else "Ya estaba cancelada.",
                "programacion": ProgramacionQuimioterapiaSerializer(cancelada).data,
            }
        )


class RecordatorioProgramacionAPIView(APIView):
    permission_classes = (EsAdministrador,)

    def patch(self, request, programacion_id):
        programacion = _programaciones_base().filter(pk=programacion_id).first()
        if not programacion:
            raise ValidationError({"programacion_id": "No existe la programación."})
        serializer = RecordatorioSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        try:
            actualizada = actualizar_recordatorio(
                programacion, usuario=request.user, **serializer.validated_data
            )
        except DjangoValidationError as error:
            raise _error_drf(error) from error
        return Response(
            {
                "detalle": "Estado del recordatorio actualizado.",
                "programacion": ProgramacionQuimioterapiaSerializer(actualizada).data,
            }
        )


class HistorialProgramacionAPIView(APIView):
    permission_classes = (EsAdministrador,)

    def get(self, request, programacion_id):
        historial = CambioProgramacionQuimioterapia.objects.filter(
            programacion_id=programacion_id
        ).select_related("realizada_por")
        return Response(
            {
                "programacion_id": str(programacion_id),
                "historial": [
                    {
                        "id": str(item.id),
                        "accion": item.accion,
                        "datos_anteriores": item.datos_anteriores,
                        "datos_nuevos": item.datos_nuevos,
                        "motivo": item.motivo,
                        "realizada_por": (
                            item.realizada_por.nombre_completo if item.realizada_por else "Sistema"
                        ),
                        "creado_en": item.creado_en,
                    }
                    for item in historial
                ],
            }
        )


def _seguro_hoja(valor):
    if isinstance(valor, str) and valor.startswith(("=", "+", "-", "@")):
        return "'" + valor
    return valor


def _filas_exportacion(fecha_desde, fecha_hasta, *, incluir_canceladas=False):
    programaciones = (
        _programaciones_base()
        .filter(fecha__range=(fecha_desde, fecha_hasta))
        .order_by("fecha", "turno", "cama")
    )
    if not incluir_canceladas:
        programaciones = programaciones.exclude(
            estado=ProgramacionQuimioterapia.Estado.CANCELADA
        )
    for programacion in programaciones:
        solicitud = programacion.solicitud
        yield [_seguro_hoja(valor) for valor in [
            solicitud.codigo_externo or "",
            solicitud.dni,
            solicitud.nombre_completo,
            solicitud.paciente.historia_clinica if solicitud.paciente else solicitud.historia_clinica_importada,
            solicitud.telefono,
            solicitud.procedencia,
            solicitud.diagnostico,
            solicitud.protocolo,
            solicitud.prioridad,
            solicitud.duracion_minutos,
            programacion.fecha.isoformat(),
            programacion.turno,
            programacion.hora_inicio.strftime("%H:%M"),
            programacion.hora_fin.strftime("%H:%M"),
            programacion.cama,
            programacion.estado,
            programacion.recordatorio_estado,
        ]]


class ExportarAgendaAPIView(APIView):
    permission_classes = (EsAdministrador,)
    CABECERAS = (
        "codigo_solicitud",
        "dni",
        "nombres_completos",
        "historia_clinica",
        "telefono",
        "procedencia",
        "diagnostico",
        "protocolo_quimioterapia",
        "prioridad",
        "duracion_minutos",
        "fecha",
        "turno",
        "hora_inicio",
        "hora_fin",
        "cama",
        "estado",
        "recordatorio",
    )

    def get(self, request):
        serializer = GenerarAgendaSerializer(
            data={
                "fecha_desde": request.query_params.get("fecha_desde"),
                "fecha_hasta": request.query_params.get("fecha_hasta") or None,
            }
        )
        serializer.is_valid(raise_exception=True)
        fecha_desde = serializer.validated_data["fecha_desde"]
        fecha_hasta = serializer.validated_data.get("fecha_hasta") or fecha_desde
        if fecha_hasta < fecha_desde:
            raise ValidationError({"fecha_hasta": "Debe ser posterior a fecha_desde."})
        incluir_canceladas = request.query_params.get("incluir_canceladas", "false").lower()
        if incluir_canceladas not in ("true", "false"):
            raise ValidationError({"incluir_canceladas": "Use true o false."})
        filas = list(
            _filas_exportacion(
                fecha_desde,
                fecha_hasta,
                incluir_canceladas=incluir_canceladas == "true",
            )
        )
        formato = request.query_params.get("formato", "xlsx").lower()
        nombre_base = f"agenda_clinica_dia_{fecha_desde.isoformat()}"
        if formato == "csv":
            contenido = StringIO()
            escritor = csv.writer(contenido)
            escritor.writerow(self.CABECERAS)
            escritor.writerows(filas)
            respuesta = HttpResponse(
                "\ufeff" + contenido.getvalue(), content_type="text/csv; charset=utf-8"
            )
            respuesta["Content-Disposition"] = f'attachment; filename="{nombre_base}.csv"'
            return respuesta
        if formato != "xlsx":
            raise ValidationError({"formato": "Use xlsx o csv."})
        libro = Workbook()
        hoja = libro.active
        hoja.title = "Agenda"
        hoja.append(self.CABECERAS)
        for fila in filas:
            hoja.append(fila)
        for celda in hoja[1]:
            celda.font = Font(bold=True, color="FFFFFF")
            celda.fill = PatternFill("solid", fgColor="009DAA")
        hoja.freeze_panes = "A2"
        contenido = BytesIO()
        libro.save(contenido)
        libro.close()
        respuesta = HttpResponse(
            contenido.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        respuesta["Content-Disposition"] = f'attachment; filename="{nombre_base}.xlsx"'
        return respuesta
